import datetime
from typing import List, Tuple

try:
	from openpyxl import Workbook
	from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
	from openpyxl.utils import get_column_letter
	exists_openpyxl = True
except Exception:
	exists_openpyxl = False

if not exists_openpyxl:
	raise SystemExit("openpyxl not installed. Please install it with: pip install openpyxl")


def auto_fit_columns(ws, min_width: int = 10, max_width: int = 40) -> None:
	for column_cells in ws.columns:
		length = 0
		col_letter = get_column_letter(column_cells[0].column)
		for cell in column_cells:
			cell_value = str(cell.value) if cell.value is not None else ""
			length = max(length, len(cell_value))
		ws.column_dimensions[col_letter].width = max(min_width, min(max_width, length + 2))


def add_headers(ws, headers: List[str]) -> None:
	header_fill = PatternFill(start_color="FFDDEBF7", end_color="FFDDEBF7", fill_type="solid")
	border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
	for idx, title in enumerate(headers, start=1):
		cell = ws.cell(row=1, column=idx, value=title)
		cell.font = Font(bold=True)
		cell.alignment = Alignment(horizontal="center", vertical="center")
		cell.fill = header_fill
		cell.border = border
	ws.freeze_panes = "A2"


def build_workbook(output_path: str) -> None:
	wb = Workbook()

	# Settings sheet
	ws_settings = wb.active
	ws_settings.title = "Settings"
	ws_settings["A1"] = "کلید"
	ws_settings["B1"] = "مقدار"
	ws_settings["A2"] = "تاریخ شروع پروژه"
	project_start = datetime.date.today()
	ws_settings["B2"] = project_start
	ws_settings["A3"] = "الگوی تعطیلی هفتگی (WORKDAY.INTL)"
	ws_settings["B3"] = "0000100"  # Friday-only weekend
	ws_settings["A5"] = "راهنما"
	ws_settings["B5"] = (
		"B2 را به تاریخ شروع دلخواه تغییر دهید. B3 الگوی تعطیلی هفتگی برای تابع WORKDAY.INTL است.\n"
		"نمونه‌ها: جمعه تعطیل '0000100'، پنجشنبه+جمعه '0001100'، شنبه+یکشنبه '1100000'، شنبه+یکشنبه (استاندارد بین‌المللی) '0000011'."
	)
	ws_settings.column_dimensions['A'].width = 35
	ws_settings.column_dimensions['B'].width = 65

	# Holidays sheet
	ws_holidays = wb.create_sheet("Holidays")
	ws_holidays["A1"] = "تاریخ تعطیلات رسمی"
	ws_holidays["B1"] = "توضیحات"
	for col in ("A", "B"):
		ws_holidays.column_dimensions[col].width = 25

	# Schedule sheet
	ws = wb.create_sheet("Schedule")
	headers = [
		"WBS",
		"نام فعالیت",
		"مدت (روز کاری)",
		"تاخیر شروع از مبدا (روز کاری)",
		"تاریخ شروع",
		"تاریخ پایان",
		"پیشنیاز (توضیحاتی)",
		"گروه/فصل",
		"یادداشت",
	]
	add_headers(ws, headers)

	# Data: (WBS, Name, Duration, StartOffset, Predecessors, Group)
	tasks: List[Tuple[str, str, int, int, str, str]] = [
		("1", "مدیریت و تجهیز کارگاه", 0, 0, "", "خلاصه"),
		("1.1", "تجهیز کارگاه", 5, 0, "شروع پروژه", "مدیریت و تجهیز"),
		("2", "عملیات تخریب", 0, 0, "", "خلاصه"),
		("2.1", "عملیات تخریب", 8, 5, "پس از تجهیز کارگاه", "عملیات تخریب"),
		("3", "عملیات خاکی", 0, 0, "", "خلاصه"),
		("3.1", "عملیات خاکی با ماشین", 15, 13, "پس از تخریب", "عملیات خاکی"),
		("3.2", "عملیات خاکی با دست", 8, 13, "همزمان با خاکی با ماشین", "عملیات خاکی"),
		("3.3", "تسطیح و رگلاژ", 5, 28, "پس از اتمام عملیات خاکی", "عملیات خاکی"),
		("4", "سازه بتنی", 0, 0, "", "خلاصه"),
		("4.1", "قالب بندی و چوب بست", 12, 33, "پس از رگلاژ", "سازه بتنی"),
		("4.2", "کارهای فولادی با میلگرد", 10, 45, "پس از قالب بندی", "سازه بتنی"),
		("4.3", "بتن درجا", 15, 55, "پس از آرماتوربندی", "سازه بتنی"),
		("4.4", "بتن پیش ساخته", 10, 45, "در صورت نیاز، هم‌پوشان با بتن درجا", "سازه بتنی"),
		("5", "سازه فلزی سبک", 0, 0, "", "خلاصه"),
		("5.1", "کارهای فولادی سبک", 10, 70, "پس از بتن درجا", "سازه فلزی سبک"),
		("6", "معماری", 0, 0, "", "خلاصه"),
		("6.1", "عملیات بنایی با سنگ", 10, 70, "پس از بتن درجا", "معماری"),
		("6.2", "اندود و بندکشی", 15, 80, "پس از بنایی", "معماری"),
		("7", "راه و محوطه سازی", 0, 0, "", "خلاصه"),
		("7.1", "زیر اساس", 8, 33, "پس از رگلاژ", "راه و محوطه سازی"),
		("7.2", "اساس", 8, 41, "پس از زیر اساس", "راه و محوطه سازی"),
		("7.3", "آسفالت", 8, 49, "پس از اساس", "راه و محوطه سازی"),
		("8", "حمل و نقل و متفرقه", 0, 0, "", "خلاصه"),
		("8.1", "حمل و نقل", 5, 33, "حین عملیات راهسازی", "پشتیبانی"),
		("8.2", "متفرقه", 5, 90, "نهایی سازی", "پشتیبانی"),
		("9", "آزمایش، تحویل و برچیدن", 0, 0, "", "خلاصه"),
		("9.1", "آزمایش و تحویل", 6, 95, "پس از اتمام کارهای اصلی", "تحویل"),
		("9.2", "برچیدن کارگاه", 3, 101, "پس از تحویل", "تحویل"),
	]

	date_style = None
	# Excel formulas for start and finish based on offsets and durations
	for idx, (wbs, name, dur, offset, preds, group) in enumerate(tasks, start=2):
		ws.cell(row=idx, column=1, value=wbs)
		ws.cell(row=idx, column=2, value=name)
		ws.cell(row=idx, column=3, value=dur)
		ws.cell(row=idx, column=4, value=offset)
		start_cell = ws.cell(row=idx, column=5)
		finish_cell = ws.cell(row=idx, column=6)
		start_cell.value = f"=WORKDAY.INTL(Settings!$B$2, D{idx}, Settings!$B$3, Holidays!$A$2:$A$1000)"
		finish_cell.value = f"=WORKDAY.INTL(E{idx}, C{idx}-1, Settings!$B$3, Holidays!$A$2:$A$1000)"
		ws.cell(row=idx, column=7, value=preds)
		ws.cell(row=idx, column=8, value=group)
		ws.cell(row=idx, column=9, value="")

	# Project end milestone aligned to 6 months from start
	idx_end = len(tasks) + 2
	ws.cell(row=idx_end, column=1, value="10")
	ws.cell(row=idx_end, column=2, value="پایان پروژه (۶ ماه)")
	ws.cell(row=idx_end, column=3, value=1)
	ws.cell(row=idx_end, column=4).value = f"=NETWORKDAYS.INTL(Settings!$B$2, EDATE(Settings!$B$2, 6), Settings!$B$3, Holidays!$A$2:$A$1000) - C{idx_end}"
	start_cell = ws.cell(row=idx_end, column=5)
	finish_cell = ws.cell(row=idx_end, column=6)
	start_cell.value = f"=WORKDAY.INTL(Settings!$B$2, D{idx_end}, Settings!$B$3, Holidays!$A$2:$A$1000)"
	finish_cell.value = f"=WORKDAY.INTL(E{idx_end}, C{idx_end}-1, Settings!$B$3, Holidays!$A$2:$A$1000)"
	ws.cell(row=idx_end, column=7, value="نشانگر پایان ۶ ماهه")
	ws.cell(row=idx_end, column=8, value="کنترل مدت")
	ws.cell(row=idx_end, column=9, value="")

	# Formatting
	auto_fit_columns(ws)
	for col in (5, 6):
		col_letter = get_column_letter(col)
		for cell in ws[col_letter]:
			cell.number_format = "yyyy-mm-dd"

	# Summary sheet
	ws_summary = wb.create_sheet("Summary")
	ws_summary["A1"] = "تاریخ پایان پروژه (حداکثر تاریخ پایان فعالیت‌ها)"
	ws_summary["A2"] = "=MAX(Schedule!F2:F200)"
	ws_summary["A2"].number_format = "yyyy-mm-dd"
	ws_summary["A4"] = "مدت کل (روز کاری)"
	ws_summary["B4"] = "=NETWORKDAYS.INTL(Settings!$B$2, A2, Settings!$B$3, Holidays!$A$2:$A$1000)"
	ws_summary["A5"] = "مدت کل (روز تقویمی)"
	ws_summary["B5"] = "=A2-Settings!$B$2+1"
	ws_summary["A6"] = "تقریب مدت (ماه تقویمی)"
	ws_summary["B6"] = "=(A2-Settings!$B$2)/30"
	ws_summary.column_dimensions['A'].width = 45
	ws_summary.column_dimensions['B'].width = 22

	# Save
	wb.save(output_path)


if __name__ == "__main__":
	build_workbook("/workspace/urban_township_schedule.xlsx")